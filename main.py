
import openpyxl
from urllib.parse import urlparse
def name():
    # Give the location of the file
    path = "//home//rpj//internpython//Input.xlsx"
    
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    
    sheet_obj = wb_obj.active
    m_row = sheet_obj.max_row
    
    # Loop will print all values
    # of first column
    lst_links = []
    lst_links = []

    for i in range(2, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 2)
        lst_links.append(cell_obj.value)

    lst_links = [x for x in lst_links if x is not None]

    pq= []
    for url in lst_links:
        a = urlparse(url)
        pq.append(a.path[1:-1])

    return pq

file_name = name()


for nami in file_name:
    
    with open(f'//home//rpj//internpython//URL FIILES//{nami}.txt', 'r') as f:
        textie = f.read()



    '''textie = HUNDRED  Hands down the quickest way to make a lot of money from home. Lots of students have genuinely made £100s from this technique. It's completely legal, risk-free, tax-free, and anyone over 18 in the UK can do it. (Not in the UK? Skip to no. 2).

    Matched betting works by taking advantage of free bets offered by betting sites and 'matching' them at a betting exchange. This removes the risk as you are betting both for and against an outcome.

    You can then squeeze out the free bet, which can be as much as £100. Multiply this by all the betting sites there are, and you can quite easily be in profit by a few hundred pounds.

    We walk you through how to make your first £13 profit (using a real example) in this gem of a guide to matched betting. If you know of any better way to make £40 an hour sitting at home, please let us know!

    Online paid surveys
    paid surveys

    A popular way for people to make money online is to fill out online surveys in their spare time. Research companies are always recruiting new members worldwide to answer surveys and test new products.

    For a few minutes of form filling, you can quickly make a couple of quid which is paid as cash or rewards. You can bag up to £3 ($5) for some surveys!

    A few good ones to try are: Branded Surveys, Swagbucks, Toluna, Pinecone, LifePoints, i-Say, Opinion Outpost, YouGov, Prizerebel, Marketagent, InboxPounds, Valued Opinions, The Opinion Panel, Mingle, Opinion Bureau, Maru Voice, Panel Base, Y Live, Survey Junkie.

    Update: See our full guide to the best paid survey sites!

    Searching the web
    Interested in how to make money fast by doing what you already do online? This has to be one of the easiest methods of making money online.

    Qmee.com rewards you for searching on Google, Bing, Yahoo, Amazon and eBay. Install a simple browser extension and sponsored results will show alongside your normal search results.

    qmee paid search

    Each Qmee result has a cash reward attached. If you are interested, simply click on it and collect your reward. You can also complete surveys to top up your earnings.

    The best thing is there is no minimum to cash out. Our first one was just 72p wired to our Paypal account. You also have the option to donate it to charity.

    Sign up now for free and start earning from your own searches!

    Online market trading
    Investing in the stock market isn't necessarily an easy way to make money, but it can be lucrative if you learn to do it properly. By the same token, you may suffer significant losses if you don't take it seriously.

    Today there is no need to fund the yachts of Wolf of Wall Street style stock brokers. You can do it all yourself with the help of online market trading platforms.

    Having spent many hours researching this new opportunity, I've been using the popular platform eToro.com.

    eToro has over 20 million users worldwide and offers free practice accounts. They featured in the BBC Two Traders: Millions by the Minute' and sponsor several Premier League football clubs.'''


    file1 = open('//home//rpj//internpython//stopwords.txt',mode='r')
    stopwods1 = (file1.read()).split()
    cont = textie.split()

    for i in cont:
        i.replace('?',' ')
        i.replace('!',' ')
        i.replace(',',' ')
        i.replace('.',' ')

    stopwords = []

    for j in stopwods1:
        stopwords.append(j.lower())

    content = []   #cleaned list of words

    for i in cont:
        if i.lower() not in stopwords:
            content.append(i) 

    file1.close()      

    def positive_score(textt):
        positive_count = 0
        with open("//home//rpj//internpython//positive-words.txt", "r", encoding="ISO-8859-1") as nf:
            nfu = nf.read()
            

        for k in textt:
            if k.lower() in nfu:
                positive_count=positive_count+1
        return positive_count


    def negative_score(textt):
        negative_count = 0
        with open("//home//rpj//internpython//negative-words.txt", "r", encoding="ISO-8859-1") as nf:
            nfu = nf.read()
            

        for k in textt:
            if k.lower() in nfu:
                negative_count=negative_count-1
        return negative_count*(-1)


    def av_word_length(content):
        lst = []
        
        for word in content:
            chr_count = 0
            for j in word:
                chr_count= chr_count +1
            lst.append(chr_count)
        return (sum(lst)/len(lst))

    def personal_pronouns(cont):
        pp = ['I','we','We','My','my','Ours','ours','us',"Ours's","our's"]
        countt=0
        for i in pp:
            countt = countt + cont.count(i)
        return countt

    def complex_word(content):
        lest = 0
        for i in content:
            count = 0
            if i[-1:-3] not in ['es','ed']:
                for j in i:
                    if j.lower() in ['a','e','i','o','u']:
                        count = count +1
            if count>=2:
                lest = lest + 1

        return (lest)







    polarity_score = ((positive_score(content)-negative_score(content))/(positive_score(content)+negative_score(content)))+0.000001 #polarity score




    subjectivity_score = ((positive_score(content)+negative_score(content))/len(content))+0.000001 #subjectivity score

    splitted = textie.split('.')

    no_of_sentence = len(splitted)

    total_no_of_words = len(cont)



    print('FILE NAME : \n',nami,'\n \n')

    print('positive score',positive_score(content))

    print('negative score',negative_score(content))

    print("polarity score",polarity_score)

    print("subjectivity score ",subjectivity_score)

    print("average sentence lenght",len(content)/no_of_sentence )

    print('Average Number of Words Per Sentence : ',total_no_of_words/no_of_sentence)

    print("word count : ",len(content))

    print("average word length : ",av_word_length(content))

    print("no of personal pronuns : ",personal_pronouns(cont))

    print("no of complex words : ",complex_word(content))

    print('percentage of complex words :',(complex_word(content)/len(content)*100),'%')

    print('Fog index : ',0.4*((len(content)/no_of_sentence)+(complex_word(content)/len(content))))

    print('#'*15)

    print('\n \n \n')








