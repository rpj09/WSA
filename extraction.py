# importing openpyxl module
import openpyxl
import requests
from bs4 import BeautifulSoup 
from urllib.parse import urlparse
 
# Give the location of the file
path = "//home//rpj//internpython//Input.xlsx"
 
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
sheet_obj = wb_obj.active
m_row = sheet_obj.max_row
 
# Loop will print all values
# of first column
lst_links = []

for i in range(2, m_row + 1):
    cell_obj = sheet_obj.cell(row = i, column = 2)
    lst_links.append(cell_obj.value)

lst_links = [x for x in lst_links if x is not None]

def name():
    pq= []
    for url in lst_links:
        a = urlparse(url)
        pq.append(a.path[1:-1])

    return pq

def writee():
    for url in lst_links:

        res = requests.get(url)

        html_page = res.content
        soup = BeautifulSoup(html_page, 'html.parser')

        title = soup.title
        #print(title.string)
        a = urlparse(url)
        paras = soup.find_all('p')
        with open(f'//home//rpj//internpython//URL FIILES//{a.path[1:-1]}.txt', 'w') as f:
            for i in paras:
                    f.write(i.get_text())










